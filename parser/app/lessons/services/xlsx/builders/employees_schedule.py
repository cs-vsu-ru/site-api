from typing import TypeAlias

from openpyxl import load_workbook

from app.employees.models import Employee
from app.lessons.models import Lesson

TargetFilename: TypeAlias = str


class XlsxEmployeesScheduleBuilder:
    def __init__(self):
        self.template_filename = 'employees_schedule_template.xlsx'
        self.target_filename = 'employees_schedule.xlsx'
        self.lesson_manager = Lesson.objects
        self.employees_manager = Employee.objects

    def build(self) -> TargetFilename:
        employees = self.employees_manager.order_by('name')
        lessons = self.lesson_manager.index_lessons(self.lesson_manager.all())
        wb = load_workbook(self.template_filename)
        ws = wb.active
        for index, employee in enumerate(employees):
            column = index + 2
            ws[1][column].value = employee.name
            for weekday in range(6):
                for number in range(8):
                    numerator_row = self._get_numerator_row(weekday, number)
                    denominator_row = numerator_row + 1
                    numerator_lesson = lessons[(employee, weekday, number, False)]
                    denominator_lesson = lessons[(employee, weekday, number, True)]
                    ws[numerator_row][column].value = self._get_cell(numerator_lesson)
                    ws[denominator_row][column].value = self._get_cell(
                        denominator_lesson
                    )
                    if self._check_merge(numerator_lesson, denominator_lesson):
                        ws.merge_cells(
                            start_row=numerator_row,
                            start_column=column + 1,
                            end_row=denominator_row,
                            end_column=column + 1,
                        )
        wb.save(self.target_filename)
        return self.target_filename

    def _get_numerator_row(self, weekday: int, number: int) -> int:
        return weekday * 16 + number * 2 + 2

    def _get_cell(self, lesson: Lesson) -> str:
        return f"{lesson.name} {lesson.groups} {lesson.placement}"

    def _check_merge(
        self, numerator_lesson: Lesson, denominator_lesson: Lesson
    ) -> bool:
        return (
            numerator_lesson.name == denominator_lesson.name
            and numerator_lesson.groups == denominator_lesson.groups
            and numerator_lesson.placement == denominator_lesson.placement
        )
