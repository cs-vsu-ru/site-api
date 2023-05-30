from typing import TypeAlias

from openpyxl import load_workbook

from app.employees.models import Employee
from app.lessons.models import Lesson

TargetFilename: TypeAlias = str


class XlsxScheduleEmployeeBuilder:
    def __init__(self, employee: Employee):
        self.employee = employee
        self.template_filename = 'schedule_employee_template.xlsx'
        self.base_target_filename = 'schedule_employee_%(id)s.xlsx'
        self.lesson_manager = Lesson.objects

    @property
    def target_filename(self) -> str:
        return self.base_target_filename % {'id': self.employee.id}

    def build(self) -> TargetFilename:
        lessons = self.lesson_manager.index_lessons(self.employee.lessons.all())
        wb = load_workbook(self.template_filename)
        ws = wb.active
        for number in range(8):
            for weekday in range(6):
                numerator_row = number * 2 + 2
                denominator_row = numerator_row + 1
                column = weekday + 1
                numerator_lesson = lessons[(self.employee, weekday, number, False)]
                denominator_lesson = lessons[(self.employee, weekday, number, True)]
                ws[numerator_row][column].value = self._get_cell(numerator_lesson)
                ws[denominator_row][column].value = self._get_cell(denominator_lesson)
                if (
                    numerator_lesson.name == denominator_lesson.name
                    and numerator_lesson.groups == denominator_lesson.groups
                    and numerator_lesson.placement == denominator_lesson.placement
                ):
                    ws.merge_cells(
                        start_row=numerator_row,
                        start_column=column + 1,
                        end_row=denominator_row,
                        end_column=column + 1,
                    )
        wb.save(self.target_filename)
        return self.target_filename

    def _get_cell(self, lesson: Lesson) -> str:
        return f"{lesson.name} {lesson.groups} {lesson.placement}"
