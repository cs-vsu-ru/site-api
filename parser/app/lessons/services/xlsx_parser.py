from openpyxl import load_workbook


class XlsxParser:
    def parse(self, filename: str, sheet_index: int = 0) -> list[list[str | None]]:
        workbook = load_workbook(filename=filename)
        sheet = workbook[workbook.sheetnames[sheet_index]]

        merged_cells_values = {}
        for merged_cell_range in sheet.merged_cells.ranges:
            top_left_cell_value = merged_cell_range.start_cell.value
            for cell in merged_cell_range.cells:
                merged_cells_values[cell] = top_left_cell_value

        data = []
        for i, row in enumerate(sheet.iter_rows()):
            row_data = []
            for j, cell in enumerate(row):
                try:
                    row_data.append(merged_cells_values[(i + 1, j + 1)])
                except KeyError:
                    row_data.append(cell.value)
            data.append(row_data)
        return data
